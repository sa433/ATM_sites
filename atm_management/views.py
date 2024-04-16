from django.shortcuts import render, get_object_or_404, redirect
from atm_management.models import ATMSite, State, City
from atm_management.forms import inputfile
from openpyxl import load_workbook
from django.contrib import messages


def inputdata(request):
    print("in inputdata")
    msg = ""
    data_list = ATMSite.objects.all()
    if request.method == 'POST':
        form = inputfile(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']

            if file.name.endswith('.xlsx'):
                wb = load_workbook(file)
                ws = wb.active

                data = []
                for row in ws.iter_rows(values_only=True):
                    data.append(row)
                print("data ", data)
                for row in data[1:]:
                    state, created = State.objects.get_or_create(sname=row[3])
                    city, created = City.objects.get_or_create(cname=row[4], state=state)

                    # Construct person_data list directly from row values
                    person_data = [row[5], row[6], row[7]]

                    ATMSite.objects.create(
                        bname=row[0],
                        bid=row[1],
                        baddress=row[2],
                        persondetail=person_data,  # Assign person_data directly
                        city=city,
                        state=state,
                    )

                msg = "Data added"  # Update the message
            elif file.name.endswith('.csv') or file.name.endswith('.pdf'):
                msg = "File Format Not supporting"
        else:
            print("Form errors:", form.errors)  # Print form validation errors
    else:
        form = inputfile()
    return render(request, 'upload.html', {'form': form, 'msg': msg, 'data_list': data_list})

def delete_item(request, bid):
    print("in delete_item")
    bid = get_object_or_404(ATMSite, bid=bid)
    print("bid ",bid)
    if request.method == "POST":
        bid.delete()
        return redirect('inputdata')
    return redirect('inputdata')

def update_item(request, bid):
    bid_objs = ATMSite.objects.filter(bid=bid)
    if bid_objs.exists():
        bid_obj = bid_objs.first()
    else:
        return render(request, "error.html", {"message": "ATMSite not found."})

    if request.method == "POST":
        new_bname = request.POST.get('new_bname')
        new_baddress = request.POST.get('new_baddress')
        new_persondetail = request.POST.get('new_persondetail')
        new_state = request.POST.get('new_state')
        new_city = request.POST.get('new_city')

        # Check which data point to update
        if new_bname:
            bid_obj.bname = new_bname
        elif new_baddress:
            bid_obj.baddress = new_baddress
        elif new_persondetail:
            bid_obj.persondetail = new_persondetail
        elif new_state:
            state, created = State.objects.get_or_create(sname=new_state)
            bid_obj.state = state
        elif new_city:
            state, created = State.objects.get_or_create(sname=new_state)
            city, created = City.objects.get_or_create(cname=new_city, state=state)
            bid_obj.city = city

        bid_obj.save()
        messages.success(request, 'Updated Successfully')
        return redirect('inputdata')
    else:
        context = {'bid_obj': bid_obj}
        return render(request, "update.html", context)